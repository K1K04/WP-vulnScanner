"""
WP VulnScanner — Módulo de Exportación
==========================================
Genera informes en Excel (.xlsx) con formato profesional
y permite comparar dos escaneos del mismo sitio.
"""

from __future__ import annotations

import io
import json
from typing import Any
from xml.sax.saxutils import escape as _xml_esc                                                


                                                                                

                                                   
from scanner._html_export import generate_standalone_html
generate_standalone_html = generate_standalone_html                      

def generate_excel(result: dict) -> bytes:
    """Genera un .xlsx profesional con múltiples hojas."""
    try:
        import openpyxl
        from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl es necesario para exportar a Excel: pip install openpyxl")

    wb = openpyxl.Workbook()
    if wb.active is not None:
        wb.remove(wb.active)                           


    SEV_CLR = {
        "critical": ("FF4757", "FFFFFF"),
        "high":     ("FF6B35", "FFFFFF"),
        "medium":   ("FFA502", "000000"),
        "low":      ("2ECC40", "000000"),
        "info":     ("00D4FF", "000000"),
    }

    def header_fill(color: str) -> PatternFill:
        return PatternFill("solid", fgColor=color)

    def thin_border() -> Border:
        s = Side(style="thin", color="30363D")
        return Border(left=s, right=s, top=s, bottom=s)

    def bold_font(color: str = "C9D1D9", size: int = 10) -> Font:
        return Font(bold=True, color=color, size=size, name="Consolas")

    def normal_font(color: str | None = "C9D1D9", size: int = 9) -> Font:
        if color is None:
            return Font(size=size, name="Consolas")
        return Font(color=color, size=size, name="Consolas")

    def set_col_widths(ws, widths: list[int]):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def style_header_row(ws, row: int, cols: int, fill_color: str = "161B22"):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill   = header_fill(fill_color)
            cell.font   = bold_font("00D4FF")
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    s  = result.get("summary", {})
    url = result.get("target_url", "")

                                                                                 
                               
                                                                                 
    ws1 = wb.create_sheet("Resumen Ejecutivo")
    ws1.sheet_view.showGridLines = False
    ws1.tab_color = "39FF14"

            
    ws1.merge_cells("A1:H1")
    t = ws1["A1"]
    t.value     = "WP VulnScanner — Informe de Seguridad"
    t.font      = Font(bold=True, size=16, color="39FF14", name="Consolas")
    t.fill      = header_fill("0D1117")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 36

               
    ws1.merge_cells("A2:H2")
    s2 = ws1["A2"]
    s2.value     = url
    s2.font      = Font(size=11, color="8B949E", name="Consolas")
    s2.fill      = header_fill("0D1117")
    s2.alignment = Alignment(horizontal="center")
    ws1.row_dimensions[2].height = 22

                       
    ws1.merge_cells("A4:B6")
    risk_cell = ws1["A4"]
    risk_score = result.get("risk_score", 0)
    risk_label = result.get("risk_label", "")
    risk_color = result.get("risk_color", "#8b949e").replace("#", "")
    risk_cell.value     = risk_score
    risk_cell.font      = Font(bold=True, size=36, color=risk_color, name="Consolas")
    risk_cell.fill      = header_fill("161B22")
    risk_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws1.merge_cells("C4:D4")
    ws1["C4"].value = risk_label
    ws1["C4"].font  = Font(bold=True, size=18, color=risk_color, name="Consolas")
    ws1["C4"].fill  = header_fill("161B22")
    ws1["C4"].alignment = Alignment(horizontal="left", vertical="center")

    ws1["C5"].value = "Puntuación de riesgo / 100"
    ws1["C5"].font  = normal_font("8B949E", 9)
    ws1["C5"].fill  = header_fill("161B22")

    ws1["C6"].value = f"Escaneado: {result.get('scanned_at', '')} · {result.get('duration', 0)}s"
    ws1["C6"].font  = normal_font("484F58", 8)
    ws1["C6"].fill  = header_fill("161B22")

                      
    metrics = [
        ("Vulnerabilidades totales", s.get("vulns_found", 0),     "FF4757"),
        ("Críticas",                 s.get("critical_vulns", 0),   "FF4757"),
        ("Altas",                    s.get("high_vulns", 0),       "FF6B35"),
        ("Medias",                   s.get("medium_vulns", 0),     "FFA502"),
        ("Plugins detectados",       s.get("plugins_found", 0),    "00D4FF"),
        ("Plugins desactualizados",  s.get("outdated_plugins", 0), "FFA502"),
        ("Archivos expuestos",       s.get("exposed_files", 0),    "FF6B35"),
        ("Usuarios expuestos",       s.get("users_found", 0),      "FF6B35"),
        ("Headers faltantes",        s.get("header_issues", 0),    "FFA502"),
        ("Malware/Spam",             s.get("malware_found", 0),    "FF4757"),
        ("XML-RPC activo",           1 if result.get("xmlrpc_enabled") else 0, "FF6B35"),
        ("WP desactualizado",        1 if result.get("wp_outdated") else 0,    "FFA502"),
    ]
    headers_m = ["Métrica", "Valor", "Estado"]
    for col, h in enumerate(headers_m, 1):
        c = ws1.cell(row=8, column=col)
        c.value = h
        c.font  = bold_font("00D4FF")
        c.fill  = header_fill("161B22")
        c.border = thin_border()
        c.alignment = Alignment(horizontal="center")

    for i, (label, value, color) in enumerate(metrics, 9):
        ws1.cell(row=i, column=1).value = label
        ws1.cell(row=i, column=1).font  = normal_font()
        ws1.cell(row=i, column=1).fill  = header_fill("0D1117" if i % 2 == 0 else "161B22")
        ws1.cell(row=i, column=1).border = thin_border()

        ws1.cell(row=i, column=2).value = value
        ws1.cell(row=i, column=2).font  = Font(bold=True, color=color, size=11, name="Consolas")
        ws1.cell(row=i, column=2).fill  = header_fill("0D1117" if i % 2 == 0 else "161B22")
        ws1.cell(row=i, column=2).border = thin_border()
        ws1.cell(row=i, column=2).alignment = Alignment(horizontal="center")

        estado = "⚠ ATENCIÓN" if value > 0 else "✓ OK"
        ws1.cell(row=i, column=3).value = estado
        ws1.cell(row=i, column=3).font  = Font(color=color if value > 0 else "2ECC40",
                                                size=9, name="Consolas")
        ws1.cell(row=i, column=3).fill  = header_fill("0D1117" if i % 2 == 0 else "161B22")
        ws1.cell(row=i, column=3).border = thin_border()

                  
    tech_data = [
        ("WordPress",  result.get("wp_version") or "No detectada"),
        ("WP Latest",  result.get("wp_latest_version") or "—"),
        ("Servidor",   result.get("server_info") or "Oculto"),
        ("PHP",        result.get("php_version") or "Oculta"),
        ("Motor vulns",  "WPScan API" if result.get("wpscan_api_used") else "Base offline"),
        ("Scan ID",    result.get("scan_id", "")),
    ]
    ws1.cell(row=8, column=5).value = "Info técnica"
    ws1.cell(row=8, column=5).font  = bold_font("00D4FF")
    ws1.cell(row=8, column=5).fill  = header_fill("161B22")
    ws1.cell(row=8, column=5).border = thin_border()
    ws1.cell(row=8, column=6).value = "Valor"
    ws1.cell(row=8, column=6).font  = bold_font("00D4FF")
    ws1.cell(row=8, column=6).fill  = header_fill("161B22")
    ws1.cell(row=8, column=6).border = thin_border()

    for i, (k, v) in enumerate(tech_data, 9):
        ws1.cell(row=i, column=5).value  = k
        ws1.cell(row=i, column=5).font   = normal_font("8B949E")
        ws1.cell(row=i, column=5).fill   = header_fill("0D1117" if i%2==0 else "161B22")
        ws1.cell(row=i, column=5).border = thin_border()
        ws1.cell(row=i, column=6).value  = str(v)
        ws1.cell(row=i, column=6).font   = normal_font()
        ws1.cell(row=i, column=6).fill   = header_fill("0D1117" if i%2==0 else "161B22")
        ws1.cell(row=i, column=6).border = thin_border()

    set_col_widths(ws1, [30, 12, 15, 5, 22, 25, 5, 5])

                                                                                 
                              
                                                                                 
    ws2 = wb.create_sheet("Vulnerabilidades")
    ws2.sheet_view.showGridLines = False
    ws2.tab_color = "FF4757"

    headers_v = ["#", "Severidad", "CVSS", "CVE", "Título", "Componente",
                 "Versión inst.", "Versión fix", "Tipo", "Estado"]
    for col, h in enumerate(headers_v, 1):
        c = ws2.cell(row=1, column=col)
        c.value = h
        c.font  = bold_font("00D4FF", 10)
        c.fill  = header_fill("161B22")
        c.border = thin_border()
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws2.row_dimensions[1].height = 30

    vulns = result.get("vulnerabilities", [])
    for i, v in enumerate(vulns, 2):
        if not isinstance(v, dict):
            continue
        sev = v.get("severity", "medium")
        bg, fg = SEV_CLR.get(sev, ("161B22", "C9D1D9"))
        row_data = [
            i - 1,
            sev.upper(),
            v.get("cvss_score") or "—",
            v.get("cve_id") or "—",
            v.get("title", ""),
            v.get("plugin_slug", ""),
            v.get("plugin_version") or "?",
            v.get("fixed_in") or "—",
            v.get("type", "plugin").upper(),
            f"{'✓ Fix: v' + v['fixed_in'] if v.get('fixed_in') else '⚠ Sin fix'}",
        ]
        for col, val in enumerate(row_data, 1):
            c = ws2.cell(row=i, column=col)
            c.value  = val
            c.border = thin_border()
            c.alignment = Alignment(vertical="center", wrap_text=(col == 5))
            if col == 2:                       
                c.fill = header_fill(bg)
                c.font = Font(bold=True, color=fg, size=9, name="Consolas")
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 3 and val != "—":                  
                cvss_color = "FF4757" if float(val) >= 9 else "FF6B35" if float(val) >= 7 else "FFA502"
                c.fill = header_fill("161B22" if i % 2 == 0 else "0D1117")
                c.font = Font(bold=True, color=cvss_color, size=9, name="Consolas")
            else:
                c.fill = header_fill("161B22" if i % 2 == 0 else "0D1117")
                c.font = normal_font(size=9)

    set_col_widths(ws2, [4, 10, 7, 16, 50, 22, 12, 12, 10, 20])
    ws2.auto_filter.ref = f"A1:J{len(vulns)+1}"
    ws2.freeze_panes    = "A2"

                                                                                 
                             
                                                                                 
    ws3 = wb.create_sheet("Plugins y Temas")
    ws3.sheet_view.showGridLines = False
    ws3.tab_color = "00D4FF"

    headers_p = ["Tipo", "Slug", "Versión instalada", "Versión más reciente",
                 "Desactualizado", "Detectado via", "Confianza"]
    for col, h in enumerate(headers_p, 1):
        c = ws3.cell(row=1, column=col)
        c.value = h
        c.font  = bold_font("00D4FF", 10)
        c.fill  = header_fill("161B22")
        c.border = thin_border()
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws3.row_dimensions[1].height = 30

    components = result.get("plugins", []) + result.get("themes", [])
    for i, p in enumerate(components, 2):
        if not isinstance(p, dict):
            continue
        is_out = p.get("is_outdated", False)
        row_data = [
            p.get("type", "plugin").upper(),
            p.get("slug", ""),
            p.get("version") or "Desconocida",
            p.get("latest_version") or "—",
            "⚠ SÍ" if is_out else "✓ No",
            p.get("detected_via", ""),
            f"{p.get('confidence', 0)}%",
        ]
        for col, val in enumerate(row_data, 1):
            c = ws3.cell(row=i, column=col)
            c.value  = val
            c.border = thin_border()
            c.fill   = header_fill("161B22" if i % 2 == 0 else "0D1117")
            c.alignment = Alignment(vertical="center")
            if col == 5:
                c.font = Font(color="FFA502" if is_out else "2ECC40",
                              bold=is_out, size=9, name="Consolas")
            else:
                c.font = normal_font(size=9)

    set_col_widths(ws3, [8, 30, 18, 20, 14, 25, 10])
    ws3.auto_filter.ref = f"A1:G{len(components)+1}"
    ws3.freeze_panes    = "A2"

                                                                                 
                                
                                                                                 
    ws4 = wb.create_sheet("Archivos Expuestos")
    ws4.sheet_view.showGridLines = False
    ws4.tab_color = "FF6B35"

    headers_f = ["Severidad", "Ruta", "URL completa", "Descripción", "Extra"]
    for col, h in enumerate(headers_f, 1):
        c = ws4.cell(row=1, column=col)
        c.value = h
        c.font  = bold_font("00D4FF")
        c.fill  = header_fill("161B22")
        c.border = thin_border()
        c.alignment = Alignment(horizontal="center")
    ws4.row_dimensions[1].height = 28

    exposed = result.get("exposed_files", [])
    for i, f in enumerate(exposed, 2):
        if isinstance(f, str):
            f = {"path": f, "url": f, "description": "", "severity": "high", "extra": ""}
        if not isinstance(f, dict):
            continue
        sev   = f.get("severity", "high")
        bg, fg = SEV_CLR.get(sev, ("161B22", "C9D1D9"))
        row_data = [sev.upper(), f.get("path",""), f.get("url",""),
                    f.get("description",""), f.get("extra","")]
        for col, val in enumerate(row_data, 1):
            c = ws4.cell(row=i, column=col)
            c.value  = val
            c.border = thin_border()
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if col == 1:
                c.fill = header_fill(bg)
                c.font = Font(bold=True, color=fg, size=9, name="Consolas")
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.fill = header_fill("161B22" if i%2==0 else "0D1117")
                c.font = normal_font(size=9)
                if col == 3:
                    c.font = Font(color="00D4FF", size=8, name="Consolas",
                                  underline="single")

    set_col_widths(ws4, [12, 40, 55, 45, 30])
    ws4.freeze_panes = "A2"

                                                                                 
                                
                                                                                 
    ws5 = wb.create_sheet("Usuarios y Headers")
    ws5.sheet_view.showGridLines = False
    ws5.tab_color = "FFA502"

    ws5["A1"].value = "USUARIOS ENUMERADOS"
    ws5["A1"].font  = bold_font("FF6B35", 12)
    ws5["A1"].fill  = header_fill("161B22")

    uh = ["ID", "Login", "Nombre visible", "Fuente"]
    for col, h in enumerate(uh, 1):
        c = ws5.cell(row=2, column=col)
        c.value = h
        c.font  = bold_font("00D4FF")
        c.fill  = header_fill("161B22")
        c.border = thin_border()

    users = result.get("users", [])
    for i, u in enumerate(users, 3):
        if not isinstance(u, dict):
            continue
        for col, val in enumerate([u.get("id",""), u.get("login",""),
                                    u.get("display_name",""), u.get("source","")], 1):
            c = ws5.cell(row=i, column=col)
            c.value  = val
            c.font   = normal_font("FFA502" if col==2 else None)
            c.fill   = header_fill("161B22" if i%2==0 else "0D1117")
            c.border = thin_border()

    offset = max(len(users) + 5, 8)

    ws5.cell(row=offset, column=1).value = "CABECERAS DE SEGURIDAD"
    ws5.cell(row=offset, column=1).font  = bold_font("FF6B35", 12)
    ws5.cell(row=offset, column=1).fill  = header_fill("161B22")

    hh = ["Estado", "Cabecera / Problema"]
    for col, h in enumerate(hh, 1):
        c = ws5.cell(row=offset+1, column=col)
        c.value = h
        c.font  = bold_font("00D4FF")
        c.fill  = header_fill("161B22")
        c.border = thin_border()

    row_h = offset + 2
    for h in result.get("headers_issues", []):
        ws5.cell(row=row_h, column=1).value = "✗ FALTA"
        ws5.cell(row=row_h, column=1).font  = Font(color="FF4757", bold=True, size=9, name="Consolas")
        ws5.cell(row=row_h, column=1).fill  = header_fill("0D1117")
        ws5.cell(row=row_h, column=1).border = thin_border()
        ws5.cell(row=row_h, column=2).value = h
        ws5.cell(row=row_h, column=2).font  = normal_font("FF4757")
        ws5.cell(row=row_h, column=2).fill  = header_fill("0D1117")
        ws5.cell(row=row_h, column=2).border = thin_border()
        row_h += 1

    for h in result.get("headers_ok", []):
        ws5.cell(row=row_h, column=1).value = "✓ OK"
        ws5.cell(row=row_h, column=1).font  = Font(color="2ECC40", bold=True, size=9, name="Consolas")
        ws5.cell(row=row_h, column=1).fill  = header_fill("161B22")
        ws5.cell(row=row_h, column=1).border = thin_border()
        ws5.cell(row=row_h, column=2).value = h
        ws5.cell(row=row_h, column=2).font  = normal_font("2ECC40")
        ws5.cell(row=row_h, column=2).fill  = header_fill("161B22")
        ws5.cell(row=row_h, column=2).border = thin_border()
        row_h += 1

    set_col_widths(ws5, [20, 60])

                                                                                 
                                                                         
                                                                                 
    ws7 = wb.create_sheet("Análisis Avanzado")
    ws7.tab_color = "00D4FF"
    style_header_row(ws7, 1, 1, "0D1117")
    ws7["A1"].value = "ANÁLISIS AVANZADO v5.1"
    ws7["A1"].font  = bold_font("39FF14", 13)
    ws7.row_dimensions[1].height = 22

    row7 = 3

    def ws7_section(title: str):
        nonlocal row7
        ws7.cell(row=row7, column=1).value = f"── {title} ──"
        ws7.cell(row=row7, column=1).font  = bold_font("00D4FF", 10)
        ws7.cell(row=row7, column=1).fill  = header_fill("161B22")
        row7 += 1

    def ws7_row(label: str, value, warn: bool = False):
        nonlocal row7
        c1 = ws7.cell(row=row7, column=1, value=label)
        c2 = ws7.cell(row=row7, column=2, value=str(value) if value is not None else "—")
        c1.font  = bold_font("8B949E", 9)
        c2.font  = normal_font("FF4757" if warn else "C9D1D9", 9)
        c1.fill  = header_fill("0D1117" if row7 % 2 == 0 else "161B22")
        c2.fill  = c1.fill
        c1.border = thin_border()
        c2.border = thin_border()
        row7 += 1

         
    csp = result.get("csp_analysis") or {}
    ws7_section("Content Security Policy (CSP)")
    ws7_row("Presente",       "✓ Sí" if csp.get("present") else "✗ Ausente", not csp.get("present"))
    ws7_row("Score CSP",      f"{csp.get('score', 0)}/100", csp.get("score", 100) < 50)
    ws7_row("unsafe-inline",  "⚠ Presente" if csp.get("unsafe_inline") else "✓ No", bool(csp.get("unsafe_inline")))
    ws7_row("unsafe-eval",    "⚠ Presente" if csp.get("unsafe_eval") else "✓ No",   bool(csp.get("unsafe_eval")))
    for issue in (csp.get("issues") or []):
        ws7_row("Issue", issue, warn=True)
    row7 += 1

          
    hsts = result.get("hsts_analysis") or {}
    ws7_section("HSTS (Strict-Transport-Security)")
    ws7_row("Presente",          "✓ Sí" if hsts.get("present") else "✗ Ausente", not hsts.get("present"))
    ws7_row("max-age",           hsts.get("max_age", "N/A"), not hsts.get("max_age_ok", True))
    ws7_row("max-age suficiente",  "✓ Sí" if hsts.get("max_age_ok") else "✗ No (<6 meses)", not hsts.get("max_age_ok"))
    ws7_row("includeSubDomains", "✓ Sí" if hsts.get("include_subdomains") else "✗ No", not hsts.get("include_subdomains"))
    ws7_row("preload",           "✓ Sí" if hsts.get("preload") else "No")
    for issue in (hsts.get("issues") or []):
        ws7_row("Issue", issue, warn=True)
    row7 += 1

             
    cookie_issues = result.get("cookie_issues") or []
    ws7_section(f"Seguridad de Cookies ({len(cookie_issues)} issues)")
    if cookie_issues:
        for ci in cookie_issues:
            ws7_row("⚠ Cookie issue", ci, warn=True)
    else:
        ws7_row("Estado", "✓ Todas las cookies con flags de seguridad correctos")
    row7 += 1

                
    js_threats = result.get("js_threats") or []
    ws7_section(f"Amenazas JS Externas ({len(js_threats)})")
    if js_threats:
        for jt in js_threats:
            ws7_row("⚠ Script sospechoso", jt, warn=True)
    else:
        ws7_row("Estado", "✓ No se detectaron scripts JS maliciosos externos")
    row7 += 1

                
    rep = result.get("reputation") or {}
    ws7_section("Reputación del Dominio")
    ws7_row("IP",           rep.get("ip", "N/A"))
    ws7_row("Nivel riesgo", rep.get("risk_level", "N/A"), rep.get("risk_level") not in ("clean", None, ""))
    ws7_row("Limpio",       "✓ Sí" if rep.get("clean", True) else "⚠ En listas negras", not rep.get("clean", True))
    for src in (rep.get("sources_flagged") or []):
        ws7_row("Lista negra", src, warn=True)
    row7 += 1

                
    robots = result.get("robots_analysis") or {}
    ws7_section("Análisis robots.txt")
    ws7_row("Presente",    "✓ Sí" if robots.get("present") else "No")
    ws7_row("Rutas Disallow sensibles", len(robots.get("disallowed_sensitive") or []))
    for w in (robots.get("warnings") or []):
        ws7_row("⚠ Aviso", w, warn=True)
    row7 += 1

                      
    admin = result.get("admin_protection") or {}
    ws7_section("Protección wp-admin")
    ws7_row("Accesible directamente", "⚠ Sí" if admin.get("accessible") else "✓ No", bool(admin.get("accessible")))
    ws7_row("Basic Auth",    "✓ Sí" if admin.get("basic_auth_required") else "No")
    ws7_row("403 Forbidden", "✓ Sí" if admin.get("forbidden") else "No")
    ws7_row("CAPTCHA",       "✓ Sí" if admin.get("has_captcha") else "No")
    for note in (admin.get("notes") or []):
        ws7_row("Nota", note)

    ws7.column_dimensions["A"].width = 32
    ws7.column_dimensions["B"].width = 70

                      
                                                                                 
    ws8 = wb.create_sheet("Raw JSON")
    ws8.tab_color = "484F58"
    ws8["A1"].value = json.dumps(result, indent=2, ensure_ascii=False)
    ws8["A1"].font  = Font(size=8, color="484F58", name="Consolas")
    ws8.column_dimensions["A"].width = 200

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


                                                                                

def compare_scans(scan_old: dict, scan_new: dict) -> dict:
    """
    Compara dos escaneos del mismo sitio y devuelve el diff estructurado.
    """

    def vuln_key(v: dict) -> str:
        return v.get("cve_id") or f"{v.get('plugin_slug')}:{v.get('title','')[:60]}"

    def plugin_key(p: dict) -> str:
        return p.get("slug", "")

    old_vulns = {vuln_key(v): v for v in scan_old.get("vulnerabilities", []) if isinstance(v, dict)}
    new_vulns = {vuln_key(v): v for v in scan_new.get("vulnerabilities", []) if isinstance(v, dict)}

    old_plugins = {plugin_key(p): p for p in
                   scan_old.get("plugins", []) + scan_old.get("themes", []) if isinstance(p, dict)}
    new_plugins = {plugin_key(p): p for p in
                   scan_new.get("plugins", []) + scan_new.get("themes", []) if isinstance(p, dict)}

    old_exposed = {f.get("path") if isinstance(f, dict) else f
                   for f in scan_old.get("exposed_files", [])}
    new_exposed = {f.get("path") if isinstance(f, dict) else f
                   for f in scan_new.get("exposed_files", [])}

                                           
    vulns_new = [new_vulns[k] for k in new_vulns if k not in old_vulns]
                                                 
    vulns_fixed = [old_vulns[k] for k in old_vulns if k not in new_vulns]
                                   
    vulns_persist = [new_vulns[k] for k in new_vulns if k in old_vulns]

                                            
    plugins_new     = [new_plugins[k] for k in new_plugins if k not in old_plugins]
    plugins_removed = [old_plugins[k] for k in old_plugins if k not in new_plugins]
    plugins_updated = []
    for k in new_plugins:
        if k in old_plugins:
            ov = old_plugins[k].get("version")
            nv = new_plugins[k].get("version")
            if ov and nv and ov != nv:
                plugins_updated.append({
                    "slug":        k,
                    "old_version": ov,
                    "new_version": nv,
                    "type":        new_plugins[k].get("type", "plugin"),
                })

                        
    files_new    = list(new_exposed - old_exposed)
    files_fixed  = list(old_exposed - new_exposed)

                          
    old_risk = scan_old.get("risk_score", 0)
    new_risk = scan_new.get("risk_score", 0)
    risk_delta = new_risk - old_risk

             
    old_hissues = set(scan_old.get("headers_issues", []))
    new_hissues = set(scan_new.get("headers_issues", []))
    headers_fixed = list(old_hissues - new_hissues)
    headers_new   = list(new_hissues - old_hissues)

                              
    if len(vulns_fixed) > len(vulns_new) and risk_delta < 0:
        status = "MEJORADO"
        status_color = "#2ed573"
    elif len(vulns_new) > len(vulns_fixed) or risk_delta > 10:
        status = "EMPEORADO"
        status_color = "#ff4757"
    elif risk_delta == 0 and not vulns_new and not vulns_fixed:
        status = "SIN CAMBIOS"
        status_color = "#8b949e"
    else:
        status = "CAMBIOS MENORES"
        status_color = "#ffa502"

    return {
        "status":          status,
        "status_color":    status_color,
        "risk_old":        old_risk,
        "risk_new":        new_risk,
        "risk_delta":      risk_delta,
        "risk_label_old":  scan_old.get("risk_label", ""),
        "risk_label_new":  scan_new.get("risk_label", ""),
        "scan_old_date":   scan_old.get("scanned_at", ""),
        "scan_new_date":   scan_new.get("scanned_at", ""),
        "scan_old_id":     scan_old.get("scan_id", ""),
        "scan_new_id":     scan_new.get("scan_id", ""),
        "target_url":      scan_new.get("target_url", ""),
        "vulns_new":       vulns_new,
        "vulns_fixed":     vulns_fixed,
        "vulns_persist":   vulns_persist,
        "plugins_new":     plugins_new,
        "plugins_removed": plugins_removed,
        "plugins_updated": plugins_updated,
        "files_new":       files_new,
        "files_fixed":     files_fixed,
        "headers_fixed":   headers_fixed,
        "headers_new":     headers_new,
        "wp_version_old":  scan_old.get("wp_version"),
        "wp_version_new":  scan_new.get("wp_version"),
        "summary": {
            "new_vulns":      len(vulns_new),
            "fixed_vulns":    len(vulns_fixed),
            "persist_vulns":  len(vulns_persist),
            "new_plugins":    len(plugins_new),
            "updated_plugins": len(plugins_updated),
            "new_files":      len(files_new),
            "fixed_files":    len(files_fixed),
            "headers_fixed":  len(headers_fixed),
        }
    }


                                                                               
                                                                    
                                                                               

def generate_executive_pdf(result: dict) -> bytes:
    """
    Mejora #11: Genera un PDF ejecutivo de 1-2 páginas orientado a no-técnicos.
    Usa semáforos de colores, lenguaje claro y recomendaciones accionables.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                         Table, TableStyle, HRFlowable)
        from reportlab.graphics.shapes import Drawing, Rect, String
        from reportlab.lib.enums import TA_CENTER
    except ImportError:
        raise ImportError("reportlab requerido: pip install reportlab")

    import io
    from datetime import datetime

    buffer = io.BytesIO()

                                                           
    C_RED    = colors.HexColor("#c1272d")                        
    C_ORANGE = colors.HexColor("#ff8c00")                           
    C_YELLOW = colors.HexColor("#ff9800")                            
    C_GREEN  = colors.HexColor("#1f7a34")                         
    C_BLUE   = colors.HexColor("#0066cc")                        
    C_DARK   = colors.HexColor("#1a1a1a")                                  
    C_LIGHT  = colors.HexColor("#f8f9fa")                      
    C_WHITE  = colors.white
    C_LGRAY  = colors.HexColor("#d4d4d4")                  

    def risk_color(score):
        if score >= 70: return C_RED
        if score >= 45: return C_ORANGE
        if score >= 20: return C_YELLOW
        return C_GREEN

    def sev_color(sev):
        return {
            "critical": C_RED, "high": C_ORANGE,
            "medium": C_YELLOW, "low": C_GREEN, "info": C_BLUE,
        }.get(str(sev).lower(), C_LGRAY)

    def sev_emoji(sev):
        return {
            "critical": "🔴 CRÍTICO", "high": "🟠 ALTO",
            "medium": "🟡 MEDIO", "low": "🟢 BAJO", "info": "🔵 INFO",
        }.get(str(sev).lower(), str(sev).upper())

    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    W = A4[0] - 4*cm                    

    def style(name="Normal", **kwargs):
        s = ParagraphStyle(name, parent=styles["Normal"])
        for k, v in kwargs.items():
            setattr(s, k, v)
        return s

    S_TITLE    = style("Title",    fontSize=24, textColor=C_DARK, spaceAfter=6, leading=30, fontName="Helvetica-Bold")
    S_SUBTITLE = style("Subtitle", fontSize=11, textColor=C_LGRAY, spaceAfter=10, leading=14)
    S_HEADING  = style("Heading",  fontSize=13, textColor=C_DARK, spaceBefore=14, spaceAfter=8,
                       fontName="Helvetica-Bold", borderPad=3)
    S_SMALL    = style("Small",    fontSize=8.5, textColor=colors.HexColor("#666666"), spaceAfter=3)

    story = []
    s = result.get("summary", {})
    score = result.get("risk_score", 0)
    rc    = risk_color(score)
    url   = result.get("target_url", "")
    now   = datetime.now().strftime("%d/%m/%Y a las %H:%M")
    label = result.get("risk_label", "")

                                                                               
    story.append(Paragraph("Informe Ejecutivo de Seguridad Web", S_TITLE))
    story.append(Paragraph(f"{url}", S_SUBTITLE))
    story.append(HRFlowable(width=W, thickness=2, color=C_BLUE, spaceAfter=6))
    story.append(Paragraph(f"Generado el {now} · Scan ID: {result.get('scan_id','N/A')}", S_SMALL))
    story.append(Spacer(1, 6))

                                                                              
                                                                       
    gauge_w, gauge_h = W, 3.2 * cm
    gauge = Drawing(gauge_w, gauge_h)

                         
    gauge_bg = Rect(0, 0, gauge_w, gauge_h)
    gauge_bg.fillColor = C_LIGHT
    gauge_bg.strokeColor = C_LGRAY
    gauge_bg.strokeWidth = 0.5
    gauge.add(gauge_bg)
                           
    bar_x, bar_y, bar_h = 0.4*cm, 0.35*cm, 0.5*cm
    bar_w = gauge_w - 5*cm
    gauge_bar_bg = Rect(bar_x, bar_y, bar_w, bar_h)
    gauge_bar_bg.fillColor = C_LGRAY
    gauge_bar_bg.strokeWidth = 0
    gauge.add(gauge_bar_bg)
                                  
    gauge_bar_fg = Rect(bar_x, bar_y, bar_w * score / 100, bar_h)
    gauge_bar_fg.fillColor = rc
    gauge_bar_fg.strokeWidth = 0
    gauge.add(gauge_bar_fg)
                   
    gauge.add(String(0.5*cm, 1.4*cm, str(score),
                     fontName="Helvetica-Bold", fontSize=38, fillColor=rc))
    gauge.add(String(0.5*cm + len(str(score))*0.6*cm, 2.1*cm, "/100",
                     fontName="Helvetica", fontSize=12, fillColor=C_LGRAY))
                        
    gauge.add(String(0.5*cm, 0.98*cm, label,
                     fontName="Helvetica-Bold", fontSize=14, fillColor=rc))
                                                     
    scale_labels = [("BAJO", 0), ("MEDIO", 0.2), ("ALTO", 0.45), ("CRÍTICO", 0.7)]
    scale_colors = [C_GREEN, C_YELLOW, C_ORANGE, C_RED]
    for i, ((txt, pct), scl_c) in enumerate(zip(scale_labels, scale_colors)):
        x_pos = bar_x + bar_w * pct
        gauge.add(String(x_pos + 2, bar_y + bar_h + 3, txt,
                         fontName="Helvetica", fontSize=6.5, fillColor=scl_c))
                                     
    stats = [
        (f"{s.get('critical_vulns',0)} críticas", C_RED),
        (f"{s.get('high_vulns',0)} altas", C_ORANGE),
        (f"{s.get('vulns_found',0)} vulns total", C_DARK),
        (f"{s.get('plugins_found',0)} plugins", C_LGRAY),
    ]
    rx = gauge_w - 4.5*cm
    for i, (txt, col) in enumerate(stats):
        gauge.add(String(rx, gauge_h - 0.65*cm - i*0.65*cm, f"● {txt}",
                         fontName="Helvetica", fontSize=9, fillColor=col))
    story.append(gauge)
    story.append(Spacer(1, 8))

                                                                                
    story.append(Paragraph("Resumen de Hallazgos", S_HEADING))

                                                
    metrics = [
        ("Vulnerabilidades críticas",
         s.get("critical_vulns", 0), C_RED,
         "Fallos que permiten a atacantes tomar control del sitio o robar datos"),
        ("Vulnerabilidades altas",
         s.get("high_vulns", 0), C_ORANGE,
         "Problemas graves que pueden comprometer la integridad del sitio"),
        ("Vulnerabilidades medias",
         s.get("medium_vulns", 0), C_YELLOW,
         "Deficiencias que facilitan ataques con menor impacto"),
        ("Archivos sensibles expuestos",
         s.get("exposed_files", 0), C_ORANGE if s.get("exposed_files",0) > 0 else C_GREEN,
         "Ficheros internos accesibles desde Internet sin autenticación"),
        ("Usuarios enumerables",
         s.get("users_found", 0), C_ORANGE if s.get("users_found",0) > 0 else C_GREEN,
         "Nombres de usuario visibles públicamente — facilita ataques de fuerza bruta"),
        ("Malware / SEO Spam",
         s.get("malware_found", 0), C_RED if s.get("malware_found",0) > 0 else C_GREEN,
         "Código malicioso o spam detectado en el sitio"),
        ("Cabeceras de seguridad faltantes",
         s.get("header_issues", 0), C_YELLOW if s.get("header_issues",0) > 3 else C_GREEN,
         "Protecciones estándar de navegador no configuradas"),
        ("Plugins desactualizados",
         s.get("outdated_plugins", 0), C_ORANGE if s.get("outdated_plugins",0) > 0 else C_GREEN,
         "Componentes con versiones antiguas que pueden tener vulnerabilidades conocidas"),
    ]

                                                  
    rep = result.get("reputation")
    if rep and not rep.get("clean", True):
        metrics.append((
            "Reputación del dominio",
            len(rep.get("sources_flagged", [])), C_RED,
            f"El dominio está en listas negras: {', '.join(rep.get('sources_flagged', []))}"
        ))

                                        
    js_t = result.get("js_threats", [])
    if js_t:
        metrics.append((
            "Scripts JS maliciosos externos",
            len(js_t), C_RED,
            "Scripts externos de terceros con comportamiento potencialmente malicioso"
        ))

                                 
    rows: list[list[Any]] = [["Área", "Estado", "Descripción"]]
    for label_m, value, color, description in metrics:
        if value == 0:
            estado = Paragraph("✅ Sin problemas", style("ok", fontSize=9, textColor=C_GREEN))
        elif color == C_RED:
            estado = Paragraph(f"🔴 {value} detectado{'s' if value > 1 else ''}", style("bad", fontSize=9, textColor=C_RED, fontName="Helvetica-Bold"))
        elif color == C_ORANGE:
            estado = Paragraph(f"🟠 {value} detectado{'s' if value > 1 else ''}", style("warn", fontSize=9, textColor=C_ORANGE, fontName="Helvetica-Bold"))
        else:
            estado = Paragraph(f"🟡 {value} detectado{'s' if value > 1 else ''}", style("info", fontSize=9, textColor=C_YELLOW, fontName="Helvetica-Bold"))

        rows.append([
            Paragraph(label_m, style("ml", fontSize=9, fontName="Helvetica-Bold")),
            estado,
            Paragraph(description, style("md", fontSize=8, textColor=colors.HexColor("#555555"))),
        ])

    metrics_table = Table(rows, colWidths=[4.5*cm, 4*cm, W - 8.5*cm])
    ts = TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0), C_DARK),
        ("TEXTCOLOR",   (0, 0), (-1, 0), C_WHITE),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0), 10),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, C_LIGHT]),
        ("BOX",         (0, 0), (-1, -1), 0.5, "#c0c0c0"),
        ("INNERGRID",   (0, 0), (-1, -1), 0.3, "#d4d4d4"),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING",  (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ])
    metrics_table.setStyle(ts)
    story.append(metrics_table)
    story.append(Spacer(1, 15))

                                                                                
    vulns = result.get("vulnerabilities", [])
    if vulns:
        story.append(Paragraph("Principales Vulnerabilidades Detectadas", S_HEADING))
        sev_order = {"critical": 0, "high": 1, "medium": 2, "low": 3, "info": 4}
        top_vulns = sorted(vulns, key=lambda v: sev_order.get(v.get("severity", "info"), 4))[:6]

        vuln_rows: list[list[Any]] = [["Severidad", "Componente", "Descripción", "Solución"]]
        for v in top_vulns:
            fix = v.get("fixed_in", "")
            fix_text = f"Actualizar a v{fix}" if fix else "Actualizar a última versión"
            vuln_rows.append([
                Paragraph(sev_emoji(v.get("severity", "info")), style("vs", fontSize=8, fontName="Helvetica-Bold")),
                Paragraph(v.get("plugin_slug", "")[:25], style("vc", fontSize=8)),
                Paragraph(v.get("title", "")[:80], style("vd", fontSize=8)),
                Paragraph(fix_text, style("vf", fontSize=8, textColor=C_BLUE)),
            ])

        vuln_table = Table(vuln_rows, colWidths=[2.5*cm, 3*cm, 7*cm, W - 12.5*cm])
        vuln_table.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), C_DARK),
            ("TEXTCOLOR",     (0, 0), (-1, 0), C_WHITE),
            ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, 0), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, C_LIGHT]),
            ("BOX",           (0, 0), (-1, -1), 0.5, "#c0c0c0"),
            ("INNERGRID",     (0, 0), (-1, -1), 0.3, "#d4d4d4"),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",    (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING",   (0, 0), (-1, -1), 6),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
        ]))
        story.append(vuln_table)
        story.append(Spacer(1, 15))

                                                                                
    story.append(Paragraph("Plan de Acción Recomendado", S_HEADING))

    recommendations = []

    crit = s.get("critical_vulns", 0)
    high = s.get("high_vulns", 0)
    if crit > 0:
        recommendations.append(("🔴 INMEDIATO (hoy)", f"Existen {crit} vulnerabilidades críticas. "
            "Deshabilite los plugins afectados y actualícelos de inmediato. "
            "Considera colocar el sitio en mantenimiento hasta resolver."))
    if high > 0:
        recommendations.append(("🟠 URGENTE (esta semana)", f"Hay {high} vulnerabilidades altas. "
            "Actualice los componentes afectados y revise los registros de acceso."))

    outdated = s.get("outdated_plugins", 0)
    if outdated > 0:
        recommendations.append(("🟡 IMPORTANTE (este mes)",
            f"Actualice {outdated} plugins/temas desactualizados. "
            "Los plugins desactualizados son el principal vector de ataque en WordPress."))

    if result.get("xmlrpc_enabled"):
        recommendations.append(("🟡 RECOMENDADO",
            "Deshabilite XML-RPC si no lo utiliza. Permite ataques de fuerza bruta "
            "y amplificación DDoS. Use un plugin de seguridad o .htaccess."))

    if s.get("exposed_files", 0) > 0:
        recommendations.append(("🟠 URGENTE",
            f"Se encontraron {s.get('exposed_files',0)} archivos accesibles públicamente. "
            "Restrinja el acceso a archivos de configuración, logs y backups."))

    if s.get("users_found", 0) > 0:
        recommendations.append(("🟡 RECOMENDADO",
            f"Se detectaron {s.get('users_found',0)} usuarios enumerables. "
            "Configure autenticación de doble factor (2FA) y use contraseñas fuertes."))

    headers_issues = s.get("header_issues", 0)
    if headers_issues >= 5:
        recommendations.append(("🟢 MEJORA",
            f"Faltan {headers_issues} cabeceras de seguridad. "
            "Solicite a su proveedor de hosting que configure HSTS, CSP y X-Frame-Options."))

    if not recommendations:
        recommendations.append(("✅ BUEN ESTADO",
            "No se detectaron problemas críticos. Mantenga los plugins y WordPress actualizados "
            "y realice auditorías periódicas."))

    rec_rows = []
    for priority, text in recommendations:
        rec_rows.append([
            Paragraph(priority, style("rp", fontSize=9, fontName="Helvetica-Bold")),
            Paragraph(text, style("rt", fontSize=9)),
        ])

    rec_table = Table(rec_rows, colWidths=[4*cm, W - 4*cm])
    rec_table.setStyle(TableStyle([
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, C_LIGHT]),
        ("BOX",         (0, 0), (-1, -1), 0.5, C_LGRAY),
        ("INNERGRID",   (0, 0), (-1, -1), 0.25, C_LGRAY),
        ("VALIGN",      (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING",  (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(rec_table)
    story.append(Spacer(1, 12))

                                                                                
    story.append(HRFlowable(width=W, thickness=1, color=C_LGRAY, spaceAfter=4))

    tech_info = []
    if result.get("wp_version"):
        wp_latest = result.get("wp_latest_version", "")
        outdated_wp = result.get("wp_outdated", False)
        wp_status = f"v{result['wp_version']}"
        if outdated_wp and wp_latest:
            wp_status += f" → DESACTUALIZADO (última: v{wp_latest})"
        tech_info.append(f"WordPress {wp_status}")

    if result.get("ssl_info"):
        ssl = result["ssl_info"]
        if ssl.get("valid") and ssl.get("days_left") is not None:
            tech_info.append(f"SSL: válido ({ssl['days_left']} días restantes)")
        elif not ssl.get("valid"):
            tech_info.append("SSL: ⚠ certificado inválido")

    rep = result.get("reputation")
    if rep:
        rl = rep.get("risk_level", "clean")
        ip = rep.get("ip", "")
        tech_info.append(f"Reputación: {rl.upper()}{' — ' + ip if ip else ''}")

    subs = result.get("subdomains", [])
    if subs:
        alive_count = sum(1 for s in subs if s.get("alive"))
        wp_count    = sum(1 for s in subs if s.get("is_wordpress"))
        tech_info.append(f"Subdominios: {len(subs)} detectados ({alive_count} activos, {wp_count} con WordPress)")

    story.append(Paragraph(
        " &nbsp;·&nbsp; ".join(tech_info) if tech_info else "Información técnica no disponible",
        style("footer", fontSize=8, textColor=C_LGRAY)
    ))
    story.append(Spacer(1, 4))
    story.append(Paragraph(
        "Este informe fue generado automáticamente por WP VulnScanner v5.1 con fines de auditoría "
        "de seguridad. Solo debe utilizarse en sistemas con autorización expresa del propietario.",
        style("disclaimer", fontSize=7, textColor=C_LGRAY, alignment=TA_CENTER)
    ))

                                                                                     
    try:
        doc.build(story)
    except Exception as _build_err:
        import logging as _log
        _log.getLogger("wpvulnscan").error("export PDF build error: %s", _build_err)
                                                              
        buffer = io.BytesIO()
    return buffer.getvalue()


                                                                               
                                                                               
                                                                               


                                                                               
                                                 
                                                                               

def generate_progress_pdf(scan_old: dict, scan_new: dict, diff: dict) -> bytes:
    """
    Genera un PDF de una página que resume el progreso de seguridad
    entre dos escaneos del mismo sitio.

    Estructura:
      - Cabecera: URL, fechas, días transcurridos
      - Semáforo de tendencia (MEJORADO / EMPEORADO / ESTABLE)
      - Score anterior vs nuevo con delta
      - Tabla de vulnerabilidades: resueltas / nuevas / persistentes
      - Sección por severidad (críticas, altas, medias, bajas)
      - Pie legal

    Buenas prácticas:
      - Importa reportlab solo aquí → no rompe el arranque si no está instalado
      - Lanza NotImplementedError explícito si falta la dependencia
      - Usa la misma paleta de colores que el resto de exportaciones
    """
    try:
        from reportlab.lib import colors as rl_colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib.enums import TA_CENTER              
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table,
            TableStyle, HRFlowable            
        )
    except ImportError as e:
        raise NotImplementedError("reportlab no instalado") from e

    import io as _io

                                                                                 
    C = {
        "bg":      rl_colors.HexColor("#08090E"),
        "accent":  rl_colors.HexColor("#2B7FFF"),
        "green":   rl_colors.HexColor("#30B86B"),
        "red":     rl_colors.HexColor("#E5484D"),
        "orange":  rl_colors.HexColor("#F4753A"),
        "amber":   rl_colors.HexColor("#F5A31A"),
        "text":    rl_colors.HexColor("#DCE1F0"),
        "text2":   rl_colors.HexColor("#8890B0"),
        "text3":   rl_colors.HexColor("#454D6A"),
        "border":  rl_colors.HexColor("#1E2236"),
        "row_a":   rl_colors.HexColor("#0D0F18"),
        "row_b":   rl_colors.HexColor("#131622"),
        "white":   rl_colors.white,
        "black":   rl_colors.black,
    }

    SEV_COLORS = {
        "critical": C["red"],
        "high":     C["orange"],
        "medium":   C["amber"],
        "low":      C["text2"],
    }

    def sty(name="body", **kwargs) -> ParagraphStyle:
        defaults = {
            "fontName":   "Helvetica",
            "fontSize":   9,
            "textColor":  C["text"],
            "leading":    13,
            "spaceAfter": 4,
        }
        defaults.update(kwargs)
        return ParagraphStyle(name, **defaults)

                                                                                 
    buf    = _io.BytesIO()
    MARGIN = 1.8 * cm
    doc    = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=MARGIN, bottomMargin=MARGIN + 0.5 * cm)
    W = A4[0] - 2 * MARGIN
    story = []

    ps  = diff.get("progress_summary", {})
    url = scan_new.get("target_url", "—")

                                                                                 
    story.append(Paragraph(
        "INFORME DE PROGRESO DE SEGURIDAD",
        sty("h1", fontName="Helvetica-Bold", fontSize=16,
            textColor=C["accent"], spaceAfter=2, tracking=60)))
    story.append(Paragraph(
        url,
        sty("url", fontName="Helvetica-Bold", fontSize=11,
            textColor=C["text"], spaceAfter=2)))
    days = ps.get("days_between", 0)
    story.append(Paragraph(
        f"Escaneo inicial: {scan_old.get('scanned_at','—')}  →  "
        f"Escaneo final: {scan_new.get('scanned_at','—')}  ·  "
        f"{days} día{'s' if days != 1 else ''} entre análisis",
        sty("meta", fontSize=8, textColor=C["text2"], spaceAfter=8)))
    story.append(HRFlowable(
        width="100%", thickness=1, color=C["border"], spaceAfter=16))

                                                                                 
    trend       = ps.get("trend", "stable")
    trend_label = {
        "improving": "MEJORANDO",
        "worsening": "EMPEORANDO",
        "stable":    "SIN CAMBIOS",
        "mixed":     "CAMBIOS MIXTOS",
    }.get(trend, "—")
    trend_color = {
        "improving": C["green"],
        "worsening": C["red"],
        "stable":    C["text2"],
        "mixed":     C["amber"],
    }.get(trend, C["text2"])

    risk_old = diff.get("risk_old", 0)
    risk_new = diff.get("risk_new", 0)
    delta    = risk_new - risk_old
    delta_s  = (f"+{delta}" if delta > 0 else str(delta)) if delta != 0 else "0"
    delta_c  = C["red"] if delta > 0 else (C["green"] if delta < 0 else C["text2"])

    kpi_data = [
        [
            Paragraph("TENDENCIA", sty("kpi_lbl", fontSize=7, textColor=C["text3"])),
            Paragraph("SCORE ANTERIOR", sty("kpi_lbl", fontSize=7, textColor=C["text3"])),
            Paragraph("SCORE ACTUAL", sty("kpi_lbl", fontSize=7, textColor=C["text3"])),
            Paragraph("VARIACIÓN", sty("kpi_lbl", fontSize=7, textColor=C["text3"])),
        ],
        [
            Paragraph(trend_label, sty("kpi_v", fontName="Helvetica-Bold",
                                        fontSize=16, textColor=trend_color, leading=20)),
            Paragraph(str(risk_old), sty("kpi_v", fontName="Helvetica-Bold",
                                         fontSize=22, textColor=C["text"], leading=26)),
            Paragraph(str(risk_new), sty("kpi_v", fontName="Helvetica-Bold",
                                         fontSize=22, textColor=trend_color, leading=26)),
            Paragraph(delta_s, sty("kpi_v", fontName="Helvetica-Bold",
                                   fontSize=22, textColor=delta_c, leading=26)),
        ],
    ]
    kpi_table = Table(kpi_data, colWidths=[W/4]*4)
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,-1), C["row_a"]),
        ("BOX",          (0,0), (-1,-1), 0.5, C["border"]),
        ("INNERGRID",    (0,0), (-1,-1), 0.5, C["border"]),
        ("TOPPADDING",   (0,0), (-1,-1), 8),
        ("BOTTOMPADDING",(0,0), (-1,-1), 8),
        ("LEFTPADDING",  (0,0), (-1,-1), 12),
        ("RIGHTPADDING", (0,0), (-1,-1), 12),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 0.5 * cm))

                                                                                 
    story.append(Paragraph(
        "RESUMEN DE VULNERABILIDADES",
        sty("section_h", fontName="Helvetica-Bold", fontSize=9,
            textColor=C["text3"], spaceAfter=8, tracking=80)))

    vuln_summary = [
        ["", "Cantidad", "Críticas", "Altas", "Medias"],
        ["Resueltas", str(ps.get("vulns_fixed", 0)),
         str(ps.get("critical_fixed", 0)), "—", "—"],
        ["Nuevas",    str(ps.get("vulns_new", 0)),
         str(ps.get("critical_new", 0)), "—", "—"],
        ["Persistentes", str(ps.get("vulns_remaining", 0)), "—", "—", "—"],
    ]

                                            
    for row_idx, key in enumerate(["vulns_fixed", "vulns_new", "vulns_persist"], 1):
        vulns = diff.get(key, [])
        counts = {"critical": 0, "high": 0, "medium": 0, "low": 0}
        for v in vulns:
            if isinstance(v, dict):
                counts[v.get("severity", "low")] = counts.get(v.get("severity", "low"), 0) + 1
        vuln_summary[row_idx][2] = str(counts["critical"])
        vuln_summary[row_idx][3] = str(counts["high"])
        vuln_summary[row_idx][4] = str(counts["medium"])

    col_ws = [W*0.30, W*0.14, W*0.18, W*0.18, W*0.18]
    vs_table = Table(vuln_summary, colWidths=col_ws)
    vs_table.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0),  C["row_a"]),
        ("BACKGROUND",    (0,1), (-1,1),  rl_colors.HexColor("#0A1F14")),              
        ("BACKGROUND",    (0,2), (-1,2),  rl_colors.HexColor("#1F0A0A")),            
        ("BACKGROUND",    (0,3), (-1,3),  C["row_b"]),
        ("TEXTCOLOR",     (0,0), (-1,0),  C["text3"]),
        ("TEXTCOLOR",     (0,1), (0,1),   C["green"]),
        ("TEXTCOLOR",     (0,2), (0,2),   C["red"]),
        ("TEXTCOLOR",     (0,3), (0,3),   C["amber"]),
        ("FONTNAME",      (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE",      (0,0), (-1,-1), 9),
        ("BOX",           (0,0), (-1,-1), 0.5, C["border"]),
        ("INNERGRID",     (0,0), (-1,-1), 0.5, C["border"]),
        ("TOPPADDING",    (0,0), (-1,-1), 7),
        ("BOTTOMPADDING", (0,0), (-1,-1), 7),
        ("LEFTPADDING",   (0,0), (-1,-1), 10),
        ("RIGHTPADDING",  (0,0), (-1,-1), 10),
        ("ALIGN",         (1,0), (-1,-1), "CENTER"),
    ]))
    story.append(vs_table)
    story.append(Spacer(1, 0.4 * cm))

                                                                                 
    vulns_fixed = [v for v in diff.get("vulns_fixed", []) if isinstance(v, dict)]
    if vulns_fixed:
        story.append(Paragraph(
            f"VULNERABILIDADES RESUELTAS ({len(vulns_fixed)})",
            sty("sh", fontName="Helvetica-Bold", fontSize=9,
                textColor=C["green"], spaceAfter=6, tracking=60)))
        rows: list[list[Any]] = [["Severidad", "CVE", "Componente", "Descripcion"]]
        for v in vulns_fixed[:15]:
            sev    = v.get("severity", "low")
            rows.append([
                Paragraph(sev.upper(), sty("td", fontSize=8,
                                           textColor=SEV_COLORS.get(sev, C["text2"]),
                                           fontName="Helvetica-Bold")),
                Paragraph(_xml_esc(v.get("cve_id") or "—"), sty("td", fontSize=8,
                                                        textColor=C["accent"])),
                Paragraph(str(v.get("plugin_slug",""))[:28],
                          sty("td", fontSize=8, textColor=C["text2"])),
                Paragraph(_xml_esc(str(v.get("title",""))[:60]),
                          sty("td", fontSize=8, textColor=C["text"])),
            ])
        t = Table(rows, colWidths=[W*0.12, W*0.18, W*0.22, W*0.48])
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,0),  C["row_a"]),
            ("TEXTCOLOR",     (0,0), (-1,0),  C["text3"]),
            ("FONTNAME",      (0,0), (-1,0),  "Helvetica-Bold"),
            ("FONTSIZE",       (0,0), (-1,-1), 8),
            ("INNERGRID",     (0,0), (-1,-1), 0.3, C["border"]),
            ("BOX",           (0,0), (-1,-1), 0.5, C["border"]),
            ("TOPPADDING",    (0,0), (-1,-1), 5),
            ("BOTTOMPADDING", (0,0), (-1,-1), 5),
            ("LEFTPADDING",   (0,0), (-1,-1), 8),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [C["row_b"], C["row_a"]]),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.3 * cm))

                                                                                 
    vulns_new = [v for v in diff.get("vulns_new", []) if isinstance(v, dict)]
    if vulns_new:
        story.append(Paragraph(
            f"VULNERABILIDADES NUEVAS ({len(vulns_new)})",
            sty("sh2", fontName="Helvetica-Bold", fontSize=9,
                textColor=C["red"], spaceAfter=6, tracking=60)))
        rows: list[list[Any]] = [["Severidad", "CVE", "Componente", "Descripcion"]]
        for v in vulns_new[:15]:
            sev = v.get("severity", "low")
            rows.append([
                Paragraph(sev.upper(), sty("td", fontSize=8,
                                           textColor=SEV_COLORS.get(sev, C["text2"]),
                                           fontName="Helvetica-Bold")),
                Paragraph(_xml_esc(v.get("cve_id") or "—"), sty("td", fontSize=8,
                                                        textColor=C["accent"])),
                Paragraph(str(v.get("plugin_slug",""))[:28],
                          sty("td", fontSize=8, textColor=C["text2"])),
                Paragraph(_xml_esc(str(v.get("title",""))[:60]),
                          sty("td", fontSize=8, textColor=C["text"])),
            ])
        t = Table(rows, colWidths=[W*0.12, W*0.18, W*0.22, W*0.48])
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,0),  C["row_a"]),
            ("TEXTCOLOR",     (0,0), (-1,0),  C["text3"]),
            ("FONTNAME",      (0,0), (-1,0),  "Helvetica-Bold"),
            ("FONTSIZE",       (0,0), (-1,-1), 8),
            ("INNERGRID",     (0,0), (-1,-1), 0.3, C["border"]),
            ("BOX",           (0,0), (-1,-1), 0.5, C["border"]),
            ("TOPPADDING",    (0,0), (-1,-1), 5),
            ("BOTTOMPADDING", (0,0), (-1,-1), 5),
            ("LEFTPADDING",   (0,0), (-1,-1), 8),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [C["row_b"], C["row_a"]]),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.3 * cm))

                                                                                 
    other_items = []
    if ps.get("plugins_updated", 0):
        other_items.append(
            f"{ps['plugins_updated']} plugin(s) actualizados en el periodo"
        )
    if ps.get("files_fixed", 0):
        other_items.append(
            f"{ps['files_fixed']} archivo(s) sensible(s) eliminado(s)"
        )
    if ps.get("headers_fixed", 0):
        other_items.append(
            f"{ps['headers_fixed']} header(s) de seguridad configurado(s)"
        )

    if other_items:
        story.append(Paragraph(
            "OTRAS MEJORAS",
            sty("sh3", fontName="Helvetica-Bold", fontSize=9,
                textColor=C["text3"], spaceAfter=6, tracking=60)))
        for item in other_items:
            story.append(Paragraph(
                f"  · {item}",
                sty("li", fontSize=9, textColor=C["text2"], spaceAfter=3)))
        story.append(Spacer(1, 0.3 * cm))

                                                                                 
    story.append(Spacer(1, 0.3 * cm))
    story.append(HRFlowable(
        width="100%", thickness=0.4, color=C["border"], spaceAfter=6))
    story.append(Paragraph(
        "Informe de progreso generado por WP VulnScanner · Análisis externo y pasivo. "
        "El solicitante declara tener autorización expresa para realizar estas auditorías (Art. 264 CP).",
        sty("legal", fontSize=7, textColor=C["text3"], alignment=TA_CENTER)))

                                                            
    try:
        doc.build(story)
    except Exception as _build_err:
        import logging as _log
        _log.getLogger("wpvulnscan").error("compare PDF build error: %s", _build_err)
        buf = io.BytesIO()
    return buf.getvalue()


                                                                               
                        
                                                                               

def generate_standalone_html(result: dict) -> bytes:
    """Genera un informe HTML standalone con estilos inline."""
    import html as html_module
    
    target_url = result.get("target_url", "Sitio WordPress")
    risk_score = result.get("risk_score", 0)
    risk_label = result.get("risk_label", "DESCONOCIDO")
    summary = result.get("summary", {})
    vulns = result.get("vulnerabilities", [])
    plugins = result.get("plugins", [])
    users = result.get("users", [])
    exposed = result.get("exposed_files", [])
    
                                 
    if risk_score >= 80:
        risk_color = "#FF4757"
        risk_bg = "#FFE5E5"
    elif risk_score >= 60:
        risk_color = "#FF9F43"
        risk_bg = "#FFF3E0"
    elif risk_score >= 35:
        risk_color = "#FFD93D"
        risk_bg = "#FFFEF0"
    else:
        risk_color = "#2ED573"
        risk_bg = "#E8F8F5"
    
    html_content = f"""<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Informe WP VulnScanner</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, Cantarell, sans-serif;
            color: #2c3e50;
            background: #ecf0f1;
            padding: 20px;
        }}
        .container {{
            max-width: 900px;
            margin: 0 auto;
            background: white;
            border-radius: 8px;
            box-shadow: 0 2px 12px rgba(0,0,0,0.1);
            overflow: hidden;
        }}
        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 30px;
            text-align: center;
        }}
        .header h1 {{
            font-size: 28px;
            margin-bottom: 10px;
        }}
        .header p {{
            opacity: 0.9;
            font-size: 14px;
        }}
        .risk-score {{
            background: {risk_bg};
            border-left: 4px solid {risk_color};
            padding: 20px;
            margin: 20px;
            border-radius: 4px;
        }}
        .risk-score h2 {{
            color: {risk_color};
            font-size: 24px;
            margin-bottom: 10px;
        }}
        .content {{
            padding: 20px;
        }}
        .section {{
            margin-bottom: 30px;
        }}
        .section h3 {{
            background: #f8f9fa;
            padding: 12px 20px;
            border-left: 4px solid #667eea;
            margin-bottom: 15px;
            font-size: 16px;
            color: #2c3e50;
        }}
        .summary-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 15px;
            margin-bottom: 20px;
        }}
        .summary-item {{
            background: #f8f9fa;
            padding: 15px;
            border-radius: 4px;
            text-align: center;
            border-top: 3px solid #667eea;
        }}
        .summary-item .number {{
            font-size: 24px;
            font-weight: bold;
            color: #667eea;
        }}
        .summary-item .label {{
            font-size: 12px;
            color: #7f8c8d;
            margin-top: 5px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 15px 0;
            font-size: 13px;
        }}
        th {{
            background: #f8f9fa;
            padding: 10px;
            text-align: left;
            border-bottom: 2px solid #bdc3c7;
            font-weight: 600;
        }}
        td {{
            padding: 10px;
            border-bottom: 1px solid #ecf0f1;
        }}
        tr:hover {{
            background: #f8f9fa;
        }}
        .severity-critical {{ color: #e74c3c; font-weight: bold; }}
        .severity-high {{ color: #e67e22; font-weight: bold; }}
        .severity-medium {{ color: #f39c12; }}
        .severity-low {{ color: #27ae60; }}
        .outdated {{ color: #e74c3c; }}
        .updated {{ color: #27ae60; }}
        .footer {{
            background: #ecf0f1;
            padding: 20px;
            text-align: center;
            font-size: 12px;
            color: #7f8c8d;
            border-top: 1px solid #bdc3c7;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Informe de Seguridad WP VulnScanner</h1>
            <p>{html_module.escape(target_url)}</p>
        </div>
        
        <div class="risk-score">
            <h2>{risk_score}/100 — {html_module.escape(risk_label)}</h2>
            <p>Score de riesgo calculado en base a vulnerabilidades, configuración y exposición de datos.</p>
        </div>
        
        <div class="content">
            <div class="section">
                <h3>📊 Resumen del Escaneo</h3>
                <div class="summary-grid">
                    <div class="summary-item">
                        <div class="number">{summary.get('vulns_found', 0)}</div>
                        <div class="label">Vulnerabilidades</div>
                    </div>
                    <div class="summary-item">
                        <div class="number">{summary.get('critical_vulns', 0)}</div>
                        <div class="label">Críticas</div>
                    </div>
                    <div class="summary-item">
                        <div class="number">{summary.get('high_vulns', 0)}</div>
                        <div class="label">Altas</div>
                    </div>
                    <div class="summary-item">
                        <div class="number">{result.get('wp_version', 'N/A')}</div>
                        <div class="label">Versión WP</div>
                    </div>
                </div>
                <div class="summary-grid">
                    <div class="summary-item">
                        <div class="number">{summary.get('plugins_found', 0)}</div>
                        <div class="label">Plugins</div>
                    </div>
                    <div class="summary-item">
                        <div class="number">{summary.get('exposed_files', 0)}</div>
                        <div class="label">Archivos Expuestos</div>
                    </div>
                    <div class="summary-item">
                        <div class="number">{summary.get('users_found', 0)}</div>
                        <div class="label">Usuarios</div>
                    </div>
                    <div class="summary-item">
                        <div class="number">{result.get('scanned_at', 'N/A')}</div>
                        <div class="label">Fecha Escaneo</div>
                    </div>
                </div>
            </div>
            
            {f'''<div class="section">
                <h3>🚨 Vulnerabilidades Detectadas</h3>
                <table>
                    <thead>
                        <tr>
                            <th>CVE</th>
                            <th>Título</th>
                            <th>Severidad</th>
                            <th>Componente</th>
                        </tr>
                    </thead>
                    <tbody>
                        {"".join(f'''<tr>
                            <td>{html_module.escape(v.get('cve_id') or v.get('cve') or 'N/A')}</td>
                            <td>{html_module.escape(v.get('title', 'Sin título'))}</td>
                            <td><span class="severity-{v.get('severity', 'unknown').lower()}">{html_module.escape(v.get('severity', 'N/A'))}</span></td>
                            <td>{html_module.escape(v.get('plugin_slug') or v.get('component', 'N/A'))}</td>
                        </tr>''' for v in vulns[:20])}
                    </tbody>
                </table>
                {f'<p><em>Se muestran las primeras 20 vulnerabilidades. Total: {len(vulns)}</em></p>' if len(vulns) > 20 else ''}
            </div>''' if vulns else ''}
            
            {f'''<div class="section">
                <h3>🔌 Plugins Detectados</h3>
                <table>
                    <thead>
                        <tr>
                            <th>Nombre</th>
                            <th>Versión</th>
                            <th>Estado</th>
                        </tr>
                    </thead>
                    <tbody>
                        {"".join(f'''<tr>
                            <td>{html_module.escape(p.get('name', p.get('slug', 'Desconocido')))}</td>
                            <td>{html_module.escape(p.get('version', '?'))}</td>
                            <td><span class="{'outdated' if p.get('is_outdated') else 'updated'}">{'⚠️ Desactualizado' if p.get('is_outdated') else '✓ Actualizado'}</span></td>
                        </tr>''' for p in plugins[:20])}
                    </tbody>
                </table>
                {f'<p><em>Se muestran los primeros 20 plugins. Total: {len(plugins)}</em></p>' if len(plugins) > 20 else ''}
            </div>''' if plugins else ''}
            
            {f'''<div class="section">
                <h3>👥 Usuarios Encontrados</h3>
                <table>
                    <thead>
                        <tr>
                            <th>Usuario</th>
                            <th>ID</th>
                            <th>Método</th>
                        </tr>
                    </thead>
                    <tbody>
                        {"".join(f'''<tr>
                            <td>{html_module.escape(u.get('login', 'N/A'))}</td>
                            <td>{u.get('id', 'N/A')}</td>
                            <td>{html_module.escape(u.get('method', 'N/A'))}</td>
                        </tr>''' for u in users)}
                    </tbody>
                </table>
            </div>''' if users else ''}
            
            {f'''<div class="section">
                <h3>📁 Archivos Expuestos</h3>
                <ul style="list-style: none;">
                    {"".join(f'''<li style="padding: 8px 0; border-bottom: 1px solid #ecf0f1;">
                        <strong>{html_module.escape(f.get('path', 'N/A'))}</strong><br/>
                        <small>{html_module.escape(f.get('url', 'N/A'))}</small>
                    </li>''' for f in exposed[:15])}
                </ul>
                {f'<p><em>Se muestran los primeros 15 archivos. Total: {len(exposed)}</em></p>' if len(exposed) > 15 else ''}
            </div>''' if exposed else ''}
        </div>
        
        <div class="footer">
            <p>Informe generado por <strong>WP VulnScanner</strong> — Análisis pasivo de seguridad en WordPress</p>
            <p>El solicitante declara tener autorización expresa para realizar estas auditorías.</p>
        </div>
    </div>
</body>
</html>"""
    
    return html_content.encode("utf-8")


                                                                               
                 
                                                                               

def generate_markdown(result: dict) -> bytes:
    """Genera un informe en formato Markdown para wikis, blogs, Notion, Confluence, etc."""
    lines = []
    
                
    target_url = result.get("target_url", "Sitio WordPress")
    lines.append(f"# Informe de Seguridad WP VulnScanner")
    lines.append(f"\n**Objetivo:** {target_url}")
    lines.append(f"**Fecha:** {result.get('scanned_at', 'N/A')}")
    lines.append(f"**Duración:** {result.get('duration', 'N/A')} segundos")
    
                
    risk_score = result.get("risk_score", 0)
    risk_label = result.get("risk_label", "DESCONOCIDO")
    lines.append(f"\n## Risk Score: {risk_score}/100 — {risk_label}")
    
             
    summary = result.get("summary", {})
    lines.append("\n## Resumen")
    lines.append(f"- **Versión WordPress:** {result.get('wp_version', 'Desconocida')}")
    lines.append(f"- **Vulnerabilidades encontradas:** {summary.get('vulns_found', 0)}")
    lines.append(f"  - Críticas: {summary.get('critical_vulns', 0)}")
    lines.append(f"  - Altas: {summary.get('high_vulns', 0)}")
    lines.append(f"  - Medias: {summary.get('medium_vulns', 0)}")
    lines.append(f"- **Plugins detectados:** {summary.get('plugins_found', 0)}")
    lines.append(f"- **Archivos expuestos:** {summary.get('exposed_files', 0)}")
    lines.append(f"- **Usuarios encontrados:** {summary.get('users_found', 0)}")
    
                      
    vulns = result.get("vulnerabilities", [])
    if vulns:
        lines.append("\n## Vulnerabilidades")
        for v in vulns:
            lines.append(f"\n### {v.get('title', 'Sin título')}")
            lines.append(f"- **CVE:** {v.get('cve_id') or v.get('cve') or 'N/A'}")
            lines.append(f"- **Severidad:** {v.get('severity', 'N/A')}")
            lines.append(f"- **CVSS:** {v.get('cvss_score') or v.get('cvss') or 'N/A'}")
            lines.append(f"- **Componente:** {v.get('plugin_slug') or v.get('component', 'N/A')}")
            if v.get('description'):
                lines.append(f"- **Descripción:** {v.get('description', '')}")
            if v.get('fixed_in'):
                lines.append(f"- **Parche en:** {v.get('fixed_in')}")
    
             
    plugins = result.get("plugins", [])
    if plugins:
        lines.append("\n## Plugins Detectados")
        for p in plugins:
            is_outdated = " ⚠️" if p.get("is_outdated") else ""
            lines.append(f"- **{p.get('name', p.get('slug', 'Desconocido'))}** ({p.get('version', '?')}){is_outdated}")
    
              
    users = result.get("users", [])
    if users:
        lines.append("\n## Usuarios Encontrados")
        for u in users:
            lines.append(f"- {u.get('login', 'N/A')} (ID: {u.get('id', 'N/A')})")
    
                        
    exposed = result.get("exposed_files", [])
    if exposed:
        lines.append("\n## Archivos Expuestos")
        for f in exposed:
            lines.append(f"- [{f.get('path', 'N/A')}]({f.get('url', '#')})")
    
            
    lines.append("\n---")
    lines.append("*Informe generado por WP VulnScanner - https://github.com/psusec/wpvulnscan*")
    
    markdown_text = "\n".join(lines)
    return markdown_text.encode("utf-8")


                                                                               
                                                           
                                                                               

def generate_sarif(result: dict) -> bytes:
    """Genera un informe en formato SARIF para GitHub Advanced Security, GitLab SAST, etc."""
    import json
    from datetime import datetime
    
    vulns = result.get("vulnerabilities", [])
    target_url = result.get("target_url", "http://wordpress.site")
    
    results = []
    for v in vulns:
        results.append({
            "ruleId": v.get("cve_id") or v.get("cve") or "WP_UNKNOWN",
            "message": {
                "text": v.get("title", "Vulnerabilidad encontrada")
            },
            "locations": [{
                "physicalLocation": {
                    "address": {
                        "parentIndex": -1,
                        "relativeUri": target_url
                    }
                }
            }],
            "properties": {
                "severity": v.get("severity", "unknown"),
                "cvss_score": v.get("cvss_score") or v.get("cvss"),
                "component": v.get("plugin_slug") or v.get("component"),
                "description": v.get("description", ""),
                "fixed_in": v.get("fixed_in")
            }
        })
    
    sarif_report = {
        "version": "2.1.0",
        "$schema": "https://raw.githubusercontent.com/oasis-tcs/sarif-spec/master/Schemata/sarif-schema-2.1.0.json",
        "runs": [{
            "tool": {
                "driver": {
                    "name": "WP VulnScanner",
                    "version": "6.1",
                    "organization": "psusec",
                    "informationUri": "https://github.com/psusec/wpvulnscan",
                    "rules": []
                }
            },
            "results": results,
            "properties": {
                "scan_date": result.get("scanned_at", datetime.now().isoformat()),
                "risk_score": result.get("risk_score", 0),
                "target_url": target_url,
                "wordpress_version": result.get("wp_version"),
                "total_vulnerabilities": len(vulns)
            }
        }]
    }
    
    return json.dumps(sarif_report, ensure_ascii=False, indent=2).encode("utf-8")
